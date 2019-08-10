#!/usr/bin/env python3

import json
import sys

if (len(sys.argv) < 2):
  print("Usage: wowza-convery.py (filenames)") 
  quit()
else:
  print("Loading from file ", sys.argv[1])


from openpyxl import load_workbook

def convert(fn):
  print("Attempting to convert file: ", fn)
  wb = load_workbook(filename = sys.argv[1])
  ws = wb.active

  # K column contains the USB Device ID

  # L column contains the Full License Key
  device_id = ws['K2':'K30']

  license_key = ws['L2':'L30']

  map = {}

  for i, device in enumerate(device_id):
    if (device[0].value != None):
      map[device[0].value] = license_key[i][0].value

  return map

def save_json(my_map):

  with open('wowza.json', 'w') as outfile:
    json.dump(my_map, outfile)
    print("File converted to json")

for arg in sys.argv[1:]:
  save_json(convert(arg))

